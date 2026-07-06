"""Generate internal Excel — customer quotation layout plus calculation preview."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from calculator import QuoteCalculated, primary_line_items, sample_field_visible
from excel_external_generator import (
    ALT_FILL,
    BRAND,
    BRAND_FILL,
    BRAND_LIGHT_FILL,
    CENTER,
    FONT_TABLE_CELL,
    FONT_TABLE_HDR,
    LEFT,
    QuotationSheetLayout,
    TABLE_BORDER,
    WHITE_FILL,
    build_customer_quotation_sheet,
)

TOTAL_BG = PatternFill("solid", fgColor="FFF2CC")

# Gap column O, calculation panel P..Z
CALC_START = 16
CALC_END = 26
CALC_WIDTH_BUDGET = 95.0

CALC_SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
CALC_LABEL_FONT = Font(name="Calibri", size=11, bold=True, color=BRAND)
CALC_VALUE_FONT = Font(name="Calibri", size=11, color="333333")
CALC_NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="666666")
CALC_TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color=BRAND)

PREVIEW_HEADERS = [
    "Factory VND",
    "Factory USD",
    "Export/Unit",
    "Unit FOB",
    "Tooling",
    "Unit Profit",
    "Mold Profit",
    "Net After Export",
]

FMT_VND = "#,##0"
FMT_USD_4 = "$#,##0.0000"
FMT_USD_2 = "$#,##0.00"
FMT_INT = "0"
FMT_PCT = '0.##"%"'


def _preview_row_values(li) -> list[float]:
    net_after = li.line_net_profit - li.export_fee_per_unit * li.qty
    return [
        li.factory_unit_price_vnd,
        li.factory_unit_price_usd,
        li.export_fee_per_unit,
        li.unit_pricing_fob_usd,
        li.tooling_fee_usd,
        li.unit_price_profit,
        li.mold_fee_profit,
        net_after,
    ]


PREVIEW_FORMATS = [
    FMT_VND,
    FMT_USD_4,
    FMT_USD_4,
    FMT_USD_4,
    FMT_USD_2,
    FMT_USD_2,
    FMT_USD_2,
    FMT_USD_2,
]


def _money_display(val: float, places: int = 2) -> str:
    return f"${val:,.{places}f}"


def _pct_display(value: float) -> float:
    return value * 100 if value <= 1 else value


def _has_mold_fee(quote: QuoteCalculated) -> bool:
    return any(li.factory_mold_fee_vnd > 0 for li in quote.input.line_items)


def _merge_fill(
    ws, r1, c1, r2, c2, value, *, font, align=CENTER, fill=None, border=TABLE_BORDER, number_format=None,
):
    from excel_external_generator import _merge_fill as customer_merge

    customer_merge(
        ws, r1, c1, r2, c2, value,
        font=font, align=align, fill=fill, border=border, number_format=number_format,
    )


def _set_calc_grid_column_widths(ws, spans: list[tuple[int, int]], weights: list[float]) -> None:
    """Distribute preview column widths from content weights (mirrors quotation sheet)."""
    calc_cols = CALC_END - CALC_START + 1
    per_phys = [0.0] * calc_cols
    for (c1, c2), w in zip(spans, weights):
        n = c2 - c1 + 1
        share = w / n
        for col in range(c1, c2 + 1):
            idx = col - CALC_START
            if 0 <= idx < calc_cols:
                per_phys[idx] = max(per_phys[idx], share)

    total = sum(per_phys) or 1.0
    for idx, col in enumerate(range(CALC_START, CALC_END + 1)):
        width = per_phys[idx] / total * CALC_WIDTH_BUDGET
        ws.column_dimensions[get_column_letter(col)].width = max(10.0, width)
    ws.column_dimensions[get_column_letter(15)].width = 2.5


def _primary_totals(quote: QuoteCalculated) -> dict[str, float | str]:
    primary = primary_line_items(quote.line_items)
    return {
        "num_parts": quote.num_parts,
        "factory_order": sum(li.factory_unit_price_usd * li.qty for li in primary),
        "revenue": sum(li.unit_pricing_fob_usd * li.qty for li in primary),
        "export_fee": quote.total_export_fee_allocated,
        "unit_profit": sum(li.unit_price_profit for li in primary),
        "mold_profit": sum(li.mold_fee_profit for li in primary),
        "net_profit": sum(li.line_net_profit - li.export_fee_per_unit * li.qty for li in primary),
        "qty_label": " · ".join(
            f"{li.part_number or f'Part {li.item_no}'}: {li.qty:,} pcs" for li in primary
        ),
    }


def _write_markup_block(ws, quote: QuoteCalculated, end_row: int) -> None:
    """Place quote markups in the top-right, above the aligned preview table."""
    data_in = quote.input
    label_end = CALC_START + 4
    value_start = label_end + 1

    row = 1
    ws.row_dimensions[row].height = 28
    _merge_fill(
        ws, row, CALC_START, row, CALC_END,
        "INTERNAL — CALCULATION PREVIEW",
        font=CALC_SECTION_FONT, fill=BRAND_FILL, border=None,
    )
    row += 2

    _merge_fill(
        ws, row, CALC_START, row, CALC_END, "Quote Markups",
        font=CALC_LABEL_FONT, fill=BRAND_LIGHT_FILL, border=TABLE_BORDER,
    )
    row += 1

    markup_rows: list[tuple[str, float, str]] = [
        ("Unit Price Markup %", _pct_display(data_in.unit_price_markup_pct), FMT_PCT),
    ]
    if _has_mold_fee(quote):
        markup_rows.append(("Mold Fee Markup %", _pct_display(data_in.mold_fee_markup_pct), FMT_PCT))
    if sample_field_visible(quote.line_items):
        markup_rows.append(("Sample Markup %", _pct_display(data_in.sample_markup_pct), FMT_PCT))

    for label, val, fmt in markup_rows:
        if row >= end_row:
            break
        ws.row_dimensions[row].height = 22
        _merge_fill(
            ws, row, CALC_START, row, label_end, label,
            font=CALC_LABEL_FONT, align=LEFT, fill=WHITE_FILL, border=TABLE_BORDER,
        )
        _merge_fill(
            ws, row, value_start, row, CALC_END, val,
            font=CALC_VALUE_FONT, align=LEFT, fill=WHITE_FILL, border=TABLE_BORDER, number_format=fmt,
        )
        row += 1


def build_internal_calc_panel(
    ws,
    quote: QuoteCalculated,
    layout: QuotationSheetLayout,
) -> int:
    """Render markups and calculation preview aligned with the quotation line-item rows."""
    from excel_external_generator import _allocate_spans, _content_weights, _write_merged_table_row

    calc_cols = CALC_END - CALC_START + 1
    label_end = CALC_START + 4
    value_start = label_end + 1

    sample_rows = [_preview_row_values(li) for li in quote.line_items]
    preview_weights = _content_weights(PREVIEW_HEADERS, sample_rows, PREVIEW_FORMATS)
    preview_spans = _allocate_spans(preview_weights, calc_cols)
    preview_spans = [(c1 + CALC_START - 1, c2 + CALC_START - 1) for c1, c2 in preview_spans]
    _set_calc_grid_column_widths(ws, preview_spans, preview_weights)

    _write_markup_block(ws, quote, layout.table_header_row)

    _write_merged_table_row(
        ws, layout.table_header_row, PREVIEW_HEADERS, preview_spans,
        fill=BRAND_FILL, font=FONT_TABLE_HDR, min_h=24,
    )

    for idx, li in enumerate(quote.line_items):
        row = layout.table_data_start_row + idx
        fill = ALT_FILL if idx % 2 else WHITE_FILL
        _write_merged_table_row(
            ws, row, _preview_row_values(li), preview_spans,
            fill=fill, font=FONT_TABLE_CELL, formats=PREVIEW_FORMATS,
        )

    row = layout.table_data_end_row + 2
    totals = _primary_totals(quote)
    summary_rows: list[tuple[str, float | int, str | None]] = [
        ("P/N's", totals["num_parts"], FMT_INT),
        ("Factory Order (USD)", totals["factory_order"], FMT_USD_2),
        ("Quoted Revenue (USD)", totals["revenue"], FMT_USD_2),
        ("Export / Inspection Fee", totals["export_fee"], FMT_USD_2),
        ("Unit Profit", totals["unit_profit"], FMT_USD_2),
        ("Mold Profit", totals["mold_profit"], FMT_USD_2),
        ("Net Profit (1st qty / part)", totals["net_profit"], FMT_USD_2),
    ]
    _merge_fill(
        ws, row, CALC_START, row, CALC_END, "Summary (first quantity per P/N)",
        font=CALC_LABEL_FONT, fill=BRAND_LIGHT_FILL, border=TABLE_BORDER,
    )
    row += 1
    for label, val, fmt in summary_rows:
        ws.row_dimensions[row].height = 22
        _merge_fill(
            ws, row, CALC_START, row, label_end, label,
            font=CALC_LABEL_FONT, align=LEFT, fill=WHITE_FILL, border=TABLE_BORDER,
        )
        _merge_fill(
            ws, row, value_start, row, CALC_END, val,
            font=CALC_VALUE_FONT, align=LEFT, fill=WHITE_FILL, border=TABLE_BORDER, number_format=fmt,
        )
        row += 1

    ws.row_dimensions[row].height = 20
    _merge_fill(
        ws, row, CALC_START, row, CALC_END,
        totals["qty_label"],
        font=CALC_NOTE_FONT, align=LEFT, fill=None, border=None,
    )
    row += 1

    ws.row_dimensions[row].height = 24
    _merge_fill(
        ws, row, CALC_START, row, label_end, "Net Profit (Selected Scenario)",
        font=CALC_TOTAL_FONT, fill=TOTAL_BG, border=TABLE_BORDER,
    )
    _merge_fill(
        ws, row, value_start, row, CALC_END, totals["net_profit"],
        font=CALC_TOTAL_FONT, fill=TOTAL_BG, border=TABLE_BORDER, number_format=FMT_USD_2,
    )
    row += 1

    if quote.total_sample_profit > 0:
        ws.row_dimensions[row].height = 20
        _merge_fill(
            ws, row, CALC_START, row, CALC_END,
            f"Sample profit (informational, not in net profit): {_money_display(quote.total_sample_profit)}",
            font=CALC_NOTE_FONT, align=LEFT, fill=None, border=None,
        )
        row += 1

    return row


def apply_internal_sheet_page_setup(ws, last_row: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.print_area = f"A1:{get_column_letter(CALC_END)}{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins = PageMargins(left=0.18, right=0.18, top=0.22, bottom=0.22, header=0.3, footer=0.3)
    ws.sheet_view.zoomScale = 75
    ws.print_options.horizontalCentered = False


def generate_excel(quote: QuoteCalculated) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Internal Quotation"

    last_left, layout = build_customer_quotation_sheet(ws, quote)
    last_right = build_internal_calc_panel(ws, quote, layout)
    apply_internal_sheet_page_setup(ws, max(last_left, last_right))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
