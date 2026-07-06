"""Generate customer-facing Excel quotation — mirrors the PDF layout."""

from __future__ import annotations

import io
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from calculator import QuoteCalculated, effective_gage_list
from customer_quote_layout import (
    COMPANY_ADDRESS,
    COMPANY_EMAIL,
    COMPANY_NAME,
    WEIGHT_NOTE,
    customer_quote_footer_texts,
    customer_quote_lead_rows,
    customer_quote_meta,
    customer_quote_table_excel,
)

BASE_DIR = Path(__file__).resolve().parent
LOGO_PATH = BASE_DIR / "assets" / "logo.png"
SIGNATURE_PATH = BASE_DIR / "assets" / "signature.png"

# Match PDF palette
BRAND = "2B5797"
BRAND_LIGHT = "D9E1F2"
ROW_ALT = "F7F9FC"
NOTE_BG = "FFFBEA"
NOTE_BORDER = "E8C840"
TEXT_MUTED = "555555"
GRID = "B0BAC8"

BRAND_FILL = PatternFill("solid", fgColor=BRAND)
BRAND_LIGHT_FILL = PatternFill("solid", fgColor=BRAND_LIGHT)
ALT_FILL = PatternFill("solid", fgColor=ROW_ALT)
NOTE_FILL = PatternFill("solid", fgColor=NOTE_BG)
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

THIN = Side(style="thin", color=GRID)
TABLE_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
META_BORDER = Border(
    left=Side(style="thin", color=BRAND),
    right=Side(style="thin", color=BRAND),
    top=Side(style="thin", color=BRAND),
    bottom=Side(style="thin", color=BRAND),
)
NOTE_BORDER_STYLE = Border(
    left=Side(style="thin", color=NOTE_BORDER),
    right=Side(style="thin", color=NOTE_BORDER),
    top=Side(style="thin", color=NOTE_BORDER),
    bottom=Side(style="thin", color=NOTE_BORDER),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

FONT_COMPANY = Font(name="Calibri", size=11, color=TEXT_MUTED)
FONT_COMPANY_NAME = Font(name="Calibri", size=11, bold=True, color=BRAND)
FONT_BANNER = Font(name="Calibri", size=19, bold=True, color="FFFFFF")
FONT_META_LABEL = Font(name="Calibri", size=11, color=TEXT_MUTED)
FONT_META_VALUE = Font(name="Calibri", size=11, bold=True, color="000000")
FONT_TABLE_HDR = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_TABLE_CELL = Font(name="Calibri", size=11, color="000000")
FONT_TERMS = Font(name="Calibri", size=11, color=TEXT_MUTED)
FONT_TERMS_BOLD = Font(name="Calibri", size=11, bold=True, color=TEXT_MUTED)
FONT_NOTE = Font(name="Calibri", size=11, italic=True, color=TEXT_MUTED)
FONT_FOOTER_LABEL = Font(name="Calibri", size=11, bold=True, color="000000")

# Fixed wide landscape grid (wider than tall, like PDF landscape letter)
SHEET_COLS = 14
TABLE_WIDTH_BUDGET = 128.0
LEFT_COLS_END = 8   # ~58% for lead-time block
RIGHT_COLS_START = 9  # ~42% for gage list — line items & gage list both end at SHEET_COLS

# PDF signature box: 2.8" × 0.85" max (proportional scale)
SIG_MAX_W_PX = 2.8 * 96
SIG_MAX_H_PX = 0.85 * 96


@dataclass(frozen=True)
class QuotationSheetLayout:
    table_header_row: int
    table_data_start_row: int
    table_data_end_row: int


def _cell(ws, row, col, value, *, font=FONT_TABLE_CELL, align=CENTER, fill=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.alignment = align
    if fill:
        c.fill = fill
    if border is not None:
        c.border = border
    if number_format:
        c.number_format = number_format
    return c


def _merge_fill(
    ws, r1, c1, r2, c2, value, *, font, align=CENTER, fill=None, border=None, number_format=None,
):
    if r2 > r1 or c2 > c1:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    _cell(ws, r1, c1, value, font=font, align=align, fill=fill, border=border, number_format=number_format)


def _display_text(value, number_format: str | None = None) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)) and number_format:
        if number_format.startswith("$"):
            places = 4 if "0000" in number_format else 2
            return f"${value:,.{places}f}"
        if "#,##0" in number_format:
            if ".00" in number_format and value != int(value):
                return f"{value:,.2f}"
            return f"{value:,.0f}"
        if number_format == "0":
            return str(int(value))
    return str(value)


def _content_weights(headers: list[str], rows: list[list], formats: list[str | None] | None = None) -> list[float]:
    weights = []
    formats = formats or [None] * len(headers)
    for i, h in enumerate(headers):
        mx = len(h)
        for row in rows:
            if i < len(row):
                fmt = formats[i] if i < len(formats) else None
                mx = max(mx, len(_display_text(row[i], fmt)))
        weights.append(float(max(mx, 4)))
    return weights


def _allocate_spans(weights: list[float], total_cols: int) -> list[tuple[int, int]]:
    """Map logical columns to merged physical column ranges spanning 1..total_cols."""
    n = len(weights)
    if n == 0:
        return []
    if n >= total_cols:
        return [(i + 1, i + 1) for i in range(n)]

    total_w = sum(weights) or float(n)
    raw = [max(1, round(w / total_w * total_cols)) for w in weights]
    while sum(raw) < total_cols:
        raw[raw.index(max(raw))] += 1
    while sum(raw) > total_cols:
        idx = raw.index(max(raw))
        if raw[idx] <= 1:
            break
        raw[idx] -= 1

    spans: list[tuple[int, int]] = []
    col = 1
    for span in raw:
        spans.append((col, col + span - 1))
        col += span
    return spans


def _set_grid_column_widths(ws, spans: list[tuple[int, int]], weights: list[float]):
    """Distribute column widths across the physical grid from logical content weights."""
    per_phys = [0.0] * SHEET_COLS
    for (c1, c2), w in zip(spans, weights):
        n = c2 - c1 + 1
        share = w / n
        for c in range(c1, c2 + 1):
            if c <= SHEET_COLS:
                per_phys[c - 1] = max(per_phys[c - 1], share)

    total = sum(per_phys) or 1.0
    for col in range(1, SHEET_COLS + 1):
        w = per_phys[col - 1] / total * TABLE_WIDTH_BUDGET
        ws.column_dimensions[get_column_letter(col)].width = max(10.0, w)


def _span_width(ws, c1: int, c2: int) -> float:
    return sum(ws.column_dimensions[get_column_letter(c)].width or 10 for c in range(c1, c2 + 1))


def _write_merged_table_row(
    ws,
    row: int,
    values: list,
    spans: list[tuple[int, int]],
    *,
    fill,
    font,
    border=TABLE_BORDER,
    align=CENTER,
    min_h: float = 22,
    formats: list[str | None] | None = None,
):
    row_h = min_h
    formats = formats or [None] * len(values)
    for i, val in enumerate(values):
        if i >= len(spans):
            break
        c1, c2 = spans[i]
        fmt = formats[i] if i < len(formats) else None
        row_h = max(row_h, _row_height_for_text(_display_text(val, fmt), _span_width(ws, c1, c2), min_h=min_h))
        _merge_fill(
            ws, row, c1, row, c2, val,
            font=font, fill=fill, border=border, align=align, number_format=fmt,
        )
    ws.row_dimensions[row].height = row_h


def _fit_image_px(path: Path, max_w: float, max_h: float) -> tuple[int, int]:
    """Scale image to fit box preserving aspect ratio (matches PDF _fit_image)."""
    img = XLImage(str(path))
    iw, ih = img.width, img.height
    if iw <= 0 or ih <= 0:
        return int(max_w), int(max_h)
    scale = min(max_w / iw, max_h / ih)
    return max(1, int(iw * scale)), max(1, int(ih * scale))


def _row_height_for_text(text: str, col_width: float, min_h: float = 15) -> float:
    if not text:
        return min_h
    chars_per_line = max(14, int(col_width * 1.1))
    lines = 0
    for part in str(text).split("\n"):
        lines += max(1, (len(part) + chars_per_line - 1) // chars_per_line)
    return max(min_h, lines * 15 + 4)


def build_customer_quotation_sheet(ws, quote: QuoteCalculated) -> tuple[int, QuotationSheetLayout]:
    """Render the customer-facing quotation layout in columns 1..SHEET_COLS."""
    headers, table_rows, table_formats = customer_quote_table_excel(quote)
    sheet_end = SHEET_COLS
    table_weights = _content_weights(headers, table_rows, table_formats)
    table_spans = _allocate_spans(table_weights, sheet_end)

    meta = customer_quote_meta(quote)
    data_in = quote.input
    row = 1

    _set_grid_column_widths(ws, table_spans, table_weights)

    # ── Header: logo left, company right (each line ends at sheet_end) ──
    ws.row_dimensions[row].height = 26
    ws.row_dimensions[row + 1].height = 26
    ws.row_dimensions[row + 2].height = 26
    if LOGO_PATH.exists():
        logo = XLImage(str(LOGO_PATH))
        lw, lh = _fit_image_px(LOGO_PATH, 1.05 * 96, 0.68 * 96)
        logo.width = lw
        logo.height = lh
        ws.add_image(logo, "A1")
    _merge_fill(ws, row, RIGHT_COLS_START, row, sheet_end, COMPANY_NAME, font=FONT_COMPANY_NAME, align=RIGHT, border=None)
    _merge_fill(ws, row + 1, RIGHT_COLS_START, row + 1, sheet_end, COMPANY_ADDRESS, font=FONT_COMPANY, align=RIGHT, border=None)
    _merge_fill(ws, row + 2, RIGHT_COLS_START, row + 2, sheet_end, COMPANY_EMAIL, font=FONT_COMPANY, align=RIGHT, border=None)
    row += 4

    # ── QUOTATION SHEET banner ──
    ws.row_dimensions[row].height = 34
    _merge_fill(ws, row, 1, row, sheet_end, "QUOTATION SHEET", font=FONT_BANNER, fill=BRAND_FILL, border=None)
    row += 2

    # ── Inquiry / date meta (5-part PDF row) ──
    ws.row_dimensions[row].height = 26
    for c in range(1, sheet_end + 1):
        _cell(ws, row, c, None, fill=BRAND_LIGHT_FILL, border=META_BORDER)
    _merge_fill(ws, row, 1, row, 2, "1.) Inquiry No#:", font=FONT_META_LABEL, fill=BRAND_LIGHT_FILL, border=META_BORDER)
    _merge_fill(ws, row, 3, row, 4, meta["inquiry_no"], font=FONT_META_VALUE, fill=BRAND_LIGHT_FILL, border=META_BORDER)
    _cell(ws, row, 5, None, fill=BRAND_LIGHT_FILL, border=META_BORDER)
    _merge_fill(ws, row, 6, row, 8, "2.) Quotation Date:", font=FONT_META_LABEL, fill=BRAND_LIGHT_FILL, border=META_BORDER)
    _merge_fill(ws, row, 9, row, sheet_end, meta["quotation_date"], font=FONT_META_VALUE, fill=BRAND_LIGHT_FILL, border=META_BORDER)
    row += 1

    if meta["show_factory"]:
        ws.row_dimensions[row].height = 26
        for c in range(1, sheet_end + 1):
            _cell(ws, row, c, None, fill=BRAND_LIGHT_FILL, border=META_BORDER)
        _merge_fill(ws, row, 1, row, 2, "Factory:", font=FONT_META_LABEL, fill=BRAND_LIGHT_FILL, border=META_BORDER)
        _merge_fill(ws, row, 3, row, sheet_end, meta["factory_name"], font=FONT_META_VALUE, fill=BRAND_LIGHT_FILL, align=LEFT, border=META_BORDER)
        row += 1

    row += 1

    # ── Line items table (logical columns merged across the full grid) ──
    table_header_row = row
    _write_merged_table_row(
        ws, row, headers, table_spans, fill=BRAND_FILL, font=FONT_TABLE_HDR, min_h=24,
    )
    row += 1

    table_data_start_row = row
    for idx, values in enumerate(table_rows):
        fill = ALT_FILL if idx % 2 else WHITE_FILL
        _write_merged_table_row(
            ws, row, values, table_spans, fill=fill, font=FONT_TABLE_CELL, formats=table_formats,
        )
        row += 1
    table_data_end_row = row - 1

    row += 1

    # ── Lead-time (left) + gage list (right) ──
    lead_rows = customer_quote_lead_rows(data_in)
    gage_entries = effective_gage_list(data_in.gage_list)
    lead_start = row

    for label, val in lead_rows:
        ws.row_dimensions[row].height = max(22, _row_height_for_text(val, 28, min_h=22))
        _merge_fill(ws, row, 1, row, 3, label, font=FONT_TERMS_BOLD, align=LEFT, border=None)
        _merge_fill(ws, row, 4, row, LEFT_COLS_END, val, font=FONT_TERMS, align=LEFT, border=None)
        row += 1

    if gage_entries:
        g_row = lead_start
        _merge_fill(
            ws, g_row, RIGHT_COLS_START, g_row, sheet_end, "Gage List",
            font=Font(name="Calibri", size=11, bold=True, color=BRAND),
            fill=BRAND_LIGHT_FILL, border=TABLE_BORDER,
        )
        g_row += 1
        ws.row_dimensions[g_row].height = 22
        _merge_fill(ws, g_row, RIGHT_COLS_START, g_row, 10, "P/N", font=FONT_TERMS_BOLD, fill=BRAND_LIGHT_FILL, border=TABLE_BORDER)
        _merge_fill(ws, g_row, 11, g_row, sheet_end, "Gage", font=FONT_TERMS_BOLD, fill=BRAND_LIGHT_FILL, border=TABLE_BORDER)
        g_row += 1
        for entry in gage_entries:
            ws.row_dimensions[g_row].height = 22
            pn = entry["part_number"] or "—"
            _merge_fill(ws, g_row, RIGHT_COLS_START, g_row, 10, pn, font=FONT_TERMS, fill=WHITE_FILL, border=TABLE_BORDER)
            _merge_fill(
                ws, g_row, 11, g_row, sheet_end, entry["gage"],
                font=FONT_TERMS, fill=WHITE_FILL, border=TABLE_BORDER,
            )
            g_row += 1

    row = max(row, lead_start + len(lead_rows)) + 1

    # ── Weight note ──
    ws.row_dimensions[row].height = 40
    _merge_fill(ws, row, 1, row, sheet_end, WEIGHT_NOTE, font=FONT_NOTE, fill=NOTE_FILL, align=CENTER, border=NOTE_BORDER_STYLE)
    row += 2

    # ── Footer paragraphs ──
    for text in customer_quote_footer_texts(data_in):
        if not text or not str(text).strip():
            continue
        ws.row_dimensions[row].height = max(20, _row_height_for_text(text, TABLE_WIDTH_BUDGET / 2, 20))
        _merge_fill(ws, row, 1, row, sheet_end, text, font=FONT_TERMS, align=LEFT, border=None)
        row += 1

    row += 1

    # ── Signature (same proportional scale as PDF: 2.8" × 0.85" max) ──
    sig_h_px = SIG_MAX_H_PX
    if SIGNATURE_PATH.exists():
        _, sig_h_px = _fit_image_px(SIGNATURE_PATH, SIG_MAX_W_PX, SIG_MAX_H_PX)
    ws.row_dimensions[row].height = max(48, sig_h_px * 0.75 + 8)
    _cell(ws, row, 1, "Signature:", font=FONT_FOOTER_LABEL, align=LEFT, border=None)
    if SIGNATURE_PATH.exists():
        sig = XLImage(str(SIGNATURE_PATH))
        sw, sh = _fit_image_px(SIGNATURE_PATH, SIG_MAX_W_PX, SIG_MAX_H_PX)
        sig.width = sw
        sig.height = sh
        ws.add_image(sig, f"B{row}")
    sig_date_label_col = max(RIGHT_COLS_START, sheet_end - 3)
    _cell(ws, row, sig_date_label_col, "Date:", font=FONT_FOOTER_LABEL, align=RIGHT, border=None)
    _merge_fill(ws, row, sig_date_label_col + 1, row, sheet_end, meta["quotation_date"], font=FONT_META_VALUE, align=LEFT, border=None)

    layout = QuotationSheetLayout(
        table_header_row=table_header_row,
        table_data_start_row=table_data_start_row,
        table_data_end_row=table_data_end_row,
    )
    return row, layout


def apply_customer_sheet_page_setup(ws, last_row: int, sheet_end: int = SHEET_COLS) -> None:
    ws.sheet_view.showGridLines = False
    ws.print_area = f"A1:{get_column_letter(sheet_end)}{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins = PageMargins(left=0.18, right=0.18, top=0.22, bottom=0.22, header=0.3, footer=0.3)
    ws.sheet_view.zoomScale = 90
    ws.print_options.horizontalCentered = True


def generate_excel_external(quote: QuoteCalculated) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    last_row, _layout = build_customer_quotation_sheet(ws, quote)
    apply_customer_sheet_page_setup(ws, last_row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


generate_excel_customer = generate_excel_external
